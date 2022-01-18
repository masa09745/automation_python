import openpyxl
import os
import glob


def del_sheet():
  
  excel_files = glob.glob('必要データ/*.xlsx')
  select_file = ','.join(excel_files)
  joblist = openpyxl.load_workbook(select_file)
  worksheets = joblist.sheetnames

  needsheets = ['RTN', 'COA', 'EV', 'IO']

  for worksheet in worksheets:
    flag = False
    for needsheet in needsheets:
      if worksheet == needsheet:
        flag = False
        break
      
      else:
        flag = True
    if flag == True:
      sheetcount = len(joblist.sheetnames)
      if sheetcount >= 2:
        del joblist[worksheet]
  joblist.save(filename = 'Job List.xlsx')

del_sheet()

def create_orderlist():
  orderlist = openpyxl.Workbook()
  count = 3

  for i in range(1,count+1):
      sheet = orderlist.create_sheet()

  RTN_Order = orderlist["Sheet"]
  RTN_Order.title = "RTN"
  COA_Order = orderlist["Sheet1"]
  COA_Order.title = "COA"
  EV_Order = orderlist["Sheet2"]
  EV_Order.title = "EV"
  IO_Order = orderlist["Sheet3"]
  IO_Order.title = "IO"
  
  orderlist.save(filename = 'Order List.xlsx')
  

create_orderlist()

newlist = openpyxl.load_workbook('Job List.xlsx')
new_RTN = newlist['RTN']
new_COA = newlist['COA']
new_EV = newlist['EV']
new_IO = newlist['IO']

orderlist = openpyxl.load_workbook('Order List.xlsx')
order_RTN = orderlist['RTN']
order_COA = orderlist['COA']
order_EV = orderlist['EV']
order_IO = orderlist['IO']

def copy_order():
  for i in range(3, new_RTN.max_row + 1):
    order1 = new_RTN.cell(row = i, column = 3).value
    order_RTN.cell(row = i-2, column = 1, value = order1)
    order2 = new_COA.cell(row = i, column = 3).value
    order_COA.cell(row = i-2, column = 1, value = order2)
    order3 = new_EV.cell(row = i, column = 3).value
    order_EV.cell(row = i-2, column = 1, value = order3)
    order4 = new_IO.cell(row = i, column = 3).value
    order_IO.cell(row = i-2, column = 1, value = order4)
  orderlist.save(filename ='Order List.xlsx')

copy_order()
